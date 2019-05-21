import csv, re, os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill, Border, colors, Alignment
from openpyxl.cell import Cell
import unicodedata

def strip_accents(text):
    try:
        text = unicode(text, 'utf-8')
    except NameError: # unicode is a default on python 3
        pass
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore')
    text = text.decode("utf-8")
    text = text.replace(" ", "_")
    return str(text)

def unicode_err(text):
    try:
        print text
        #text = unicode(text, 'utf-8')
    except: # unicode is a default on python 3
        pass
    print text
    #text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore')
    text = text.decode("utf-8")
    return str(text)

def delete_blank_sheet(pathdir):
    wb = load_workbook(filename = pathdir)
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove_sheet(std)
        wb.save(pathdir)
    wb.close()
