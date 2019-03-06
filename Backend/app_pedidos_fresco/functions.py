import csv, re, os
import pandas as pd
import numpy as np
import unicodedata
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill, Border, colors, Alignment
from openpyxl.cell import Cell
# from difflib import SequenceMatcher
import warnings
warnings.filterwarnings("ignore", 'This pattern has match groups')

def strip_accents(text):
    try:
        text = unicode(text, 'utf-8')
    except NameError: # unicode is a default on python 3 
        pass
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore')
    text = text.decode("utf-8")
    return str(text)

def read_productos_por_proveedor(filepath):
    data = pd.read_csv(filepath, skiprows=1, decimal=',')
    # print data
    data = data [['Proveedor','descripcion','cantidad']] 
    data = data.dropna()
    return data

def read_productos(filepath):
    data = pd.read_csv(filepath, skiprows=1)
    data = data [['Productname','EnVenta']] 
    data.rename(columns={'Productname': 'descripcion'}, inplace=True)
    data = data.loc[data['EnVenta'] == 'S']
    #data = data.dropna()
    #print data
    return data

def add_formats(filepath, products):
    formats = pd.read_csv(filepath)
    products = products.join(formats.set_index('descripcion'), on='descripcion')
    return products


def add_provider_from_dict(products, provider_dict):
    products['Proveedor'] = ''
    for item in provider_dict:
        names = products['descripcion']
        names = names.str.lower()
        name = provider_dict[item]
        name = name.lower()
        prodIdx = names.str.contains(name)
        products['Proveedor'][prodIdx] = item
    products = products[products['Proveedor'] != '']
    return products

def get_provider_list (data_physical, data_online):
	providers_physical = data_physical[['Proveedor']]
	providers_physical = providers_physical.drop_duplicates(subset='Proveedor', keep='first')
	providers_online = data_online[['Proveedor']]
	providers_online = providers_online.drop_duplicates(subset='Proveedor', keep='first')

	provider_list_idx = providers_online.set_index('Proveedor').join(providers_physical.set_index('Proveedor'))
	provider_list = provider_list_idx.index.values.tolist()
	return provider_list

def merge_provider_data (data_products, data_physical, data_online, provider_list, output_path):
    df_dict = {}
    for provider in provider_list:
        #print provider
        products = data_products[data_products['Proveedor'] == provider]
        products = products[['descripcion','Formato']]

        products_physical = data_physical[data_physical['Proveedor'].str.contains(provider)]
        products_physical = products_physical[['descripcion','cantidad']]
        products_online = data_online[data_online['Proveedor'].str.contains(provider)]
        products_online = products_online[['descripcion','cantidad']]


        # strip products accents
        for i,row in products.iterrows():
            products.at[i,'descripcion'] = strip_accents(row['descripcion'])

        # strip products_physical accents
        for i,row in products_physical.iterrows():
            products_physical.at[i,'descripcion'] = strip_accents(row['descripcion'])

        # strip products_online accents
        for i,row in products_online.iterrows():
            products_online.at[i,'descripcion'] = strip_accents(row['descripcion'])

        # corr = []
        # for index, row in products.iterrows():
        #     for index2,row2 in products_online.iterrows():
        #         d1 = row['descripcion']
        #         d2 = row2['descripcion']
        #         c = similar(d1, d2)
        #         corr.append({'descripcion_1': d1, 'descripcion_2': d2, 'Corr': c})

        # for row in corr:
        #     if row["Corr"] > 0.8:
        #         print row

        products_provider = products.join(products_online.set_index('descripcion'), on='descripcion')
        # new_products = find_non_common (products_provider, products_online)
        products_provider = products_provider.join(products_physical.set_index('descripcion'), on='descripcion', lsuffix='_web', rsuffix='_tienda')
        products_provider["Excedentes"] = ""
        products_provider["Prevision"] = ""
        products_provider["Pedido"] = ""
        
        products_provider = products_provider.iloc[products_provider.descripcion.str.lower().argsort()] #sort not ignoring upper and lowercase
        products_provider = products_provider[['descripcion', 'cantidad_tienda', 'Excedentes', 'cantidad_web', 'Prevision', 'Pedido', 'Formato']]
        products_provider.columns = ['Nombre_producto', 'Venta_tienda', 'Excedentes', 'Venta_online', 'Prevision', 'Pedido', 'Formato']
        
        df_dict[provider] = products_provider
        
    return df_dict
    
def delete_blank_sheet(pathdir):
    wb = load_workbook(filename = pathdir)
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove_sheet(std)
        wb.save(pathdir)
    wb.close()

def save_excel(pathdir,data,nameSheet):
    if not os.path.exists(pathdir):
        # create workbook if not exists
        wb = Workbook()
        wb.save(filename = pathdir)


    # load workbook
    wb = load_workbook(filename = pathdir)
    if len(nameSheet) > 20:
        nameSheet = nameSheet[:20]
    ws = wb.create_sheet(title=nameSheet)
        
    
    # extract format from data
    formats = data['Formato']
    formats = formats.to_frame()

    # drop Formato column
    data = data.drop(columns="Formato")

    # add dataframe data
    rows = dataframe_to_rows(data, index=False, header=True)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # add format data
    rows = dataframe_to_rows(formats, index=False, header=True)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 12):
            ws.cell(row=r_idx, column=c_idx, value=value)


    #for row in rows:
        #ws.append(row)
        
    # add prevision formula
    for i in range(3, ws.max_row+1):
        n = '=(F%d-D%d)+C%d' % (i, i, i)
        a = ws.cell(column=5,row = i , value=n)
        

    #adjust column size
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        #ws.column_dimensions[column].width = adjusted_width

    # titles
    ws.merge_cells('B1:F1')
    ws.cell(row=1 , column=2, value='Viernes')
    ws.merge_cells('G1:K1')
    ws.cell(row=1 , column=7, value='Lunes')
    ws.merge_cells('L1:P1')
    ws.cell(row=1 , column=12, value='Organizacion en tienda de pedidos')
    ws.cell(row=2 , column=13, value='Cajas')
    ws.cell(row=2 , column=14, value='Para pedidos')
    ws.cell(row=2 , column=15, value='Para camara')
    ws.cell(row=2 , column=16, value='Para tienda')

    blueFill = PatternFill(start_color='FF4286f4',
               end_color='FF4286f4',
               fill_type='solid')
    greenFill = PatternFill(start_color='FF41f49b',
               end_color='FF41f49b',
               fill_type='solid')
    green2Fill = PatternFill(start_color='FFd4fc79',
               end_color='FFd4fc79',
               fill_type='solid')

    ws['B1'].alignment = Alignment(horizontal='center')
    ws['B1'].fill = blueFill
    ws['B2'].fill = blueFill
    ws['C2'].fill = blueFill
    ws['D2'].fill = blueFill
    ws['E2'].fill = blueFill
    ws['F2'].fill = blueFill
    ws['G1'].alignment = Alignment(horizontal='center')
    ws['G1'].fill = greenFill
    ws['G2'].fill = greenFill
    ws['H2'].fill = greenFill
    ws['I2'].fill = greenFill
    ws['J2'].fill = greenFill
    ws['K2'].fill = greenFill
    ws['L1'].alignment = Alignment(horizontal='center')
    ws['L1'].fill = green2Fill
    ws['L2'].fill = green2Fill
    ws['M2'].fill = green2Fill
    ws['N2'].fill = green2Fill
    ws['O2'].fill = green2Fill
    ws['P2'].fill = green2Fill

    #put first row in bold
    first_row = ws[1]
    for column in first_row:
        column.font = Font(bold=True,size=18)
    scnd_row = ws[2]
    for column in scnd_row:
        column.font = Font(bold=True,size=14)
            
    wb.save(pathdir)
    wb.close()  


def edit_excel(pathdir,data,nameSheet):

    print ('edit_excel function - ' + nameSheet)
    #print ('pathdir is: ', pathdir)
    #print ('name sheet is: ', nameSheet)
    # print ('data is: ' + data)
    data = data.drop(columns=['Nombre_producto'])
    wb = load_workbook(filename = pathdir)

    if len(nameSheet) > 20:
        nameSheet = nameSheet[:20]
    try:
        ws = wb[nameSheet]
        # add dataframe data
        rows = dataframe_to_rows(data, index=False, header=True)
        for r_idx, row in enumerate(rows, 2):
            for c_idx, value in enumerate(row, 7):
                ws.cell(row=r_idx, column=c_idx, value=value)
        #for row in rows:
            #ws.append(row)
            
        # add prevision formula
        for i in range(3, ws.max_row+1):
            n = '=(K%d-I%d)+H%d' % (i, i, i)
            a = ws.cell(column=10,row = i , value=n)
        
        # OJO ESTO ES LO QUE PETA
        #adjust column size 
        # for col in ws.columns:
        #     max_length = 0
        #     column = col[0].column # Get the column name
        #     for cell in col:
        #         try: # Necessary to avoid error on empty cells
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2)
        #     # ws.column_dimensions[column].width = adjusted_width
                
        wb.save(pathdir)
        wb.close()  
    except:
        print 'error en edit_excel'
        
def add_new_products_excel(pathdir, data, nameSheet):
    data = data.drop(columns=['_merge'])
    wb = load_workbook(filename = pathdir)
    try:
        ws = wb[nameSheet]
        # add dataframe data
        rows = dataframe_to_rows(data, index=False, header=False)
        for r_idx, row in enumerate(rows, ws.max_row+1):
            for c_idx, value in enumerate(row, 1):
                # print r_idx
                # print c_idx
                # print value
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(pathdir)
        wb.close()  
    except:
        pass

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

# def find_non_common (df1, df2):
#     df_all = df1.merge(df2.drop_duplicates(), on=['descripcion'], 
#                    how='left', indicator=True)
#     print df_all
#     df_all['_merge'] == 'left_only'
#     print df_all
#     return 
